"""
Budget Service — CRUD operations + Allocation logic and budget vs actual reporting.
"""

from typing import Dict, List, Any, Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, extract, func
from sqlalchemy.orm import selectinload
from app.models.property import Property
from app.models.contract import Contract, ContractStatus
from app.models.budget import Budget, BudgetCategory, BudgetRevision
from app.models.financial import Transaction, TransactionDirection
from app.schemas.budget import BudgetCreate, BudgetDuplicate


def _year_eq(col, year: int):
    """Cross-DB year filter: works for both SQLite and PostgreSQL."""
    return extract('year', col) == year


def _month_between(col, start_m: int, end_m: int):
    """Cross-DB month range filter: works for both SQLite and PostgreSQL."""
    m = extract('month', col)
    return and_(m >= start_m, m <= end_m)


def _month_eq(col, month: int):
    """Cross-DB exact month filter: works for both SQLite and PostgreSQL."""
    return extract('month', col) == month


async def _refresh_budget_totals(db: AsyncSession, budget: Budget):
    """
    Calculates execution totals for a budget and its categories in real-time.
    Also updates total_budget if auto_calculate_total is True.
    """
    if budget.is_closed and budget.frozen_distribution is not None:
        dist_keys = budget.frozen_distribution
    else:
        dist_keys = await calculate_distribution_keys(db, budget.property_id) if budget.property_id else {}
        
    all_prop_ids = list(dist_keys.keys()) + [budget.property_id]

    from app.models.financial import TransactionDirection

    # Optimized: One query for all relevant transactions in the period
    all_prop_ids_filtered = [pid for pid in all_prop_ids if pid is not None]

    # Determine the months included in the period
    period = getattr(budget, "period_type", "Mensual")
    start_m = budget.month
    if period == "Bimestral": end_m = start_m + 1
    elif period == "Trimestral": end_m = start_m + 2
    elif period == "Semestral": end_m = start_m + 5
    elif period == "Anual":
        start_m = 1
        end_m = 12
    else: end_m = start_m

    # Use EXTRACT() — compatible with both SQLite (via aiosqlite) and PostgreSQL
    query = select(Transaction.budget_category_id, func.coalesce(func.sum(Transaction.amount), 0.0)).where(
        and_(
            Transaction.direction == TransactionDirection.CREDIT.value,
            Transaction.status == "Completada",
            _year_eq(Transaction.transaction_date, budget.year),
            _month_between(Transaction.transaction_date, start_m, end_m),
            Transaction.budget_category_id != None
        )
    )

    # Handle property IDs (including NULL for general expenses if applicable)
    if None in all_prop_ids:
        query = query.where(
            (Transaction.property_id.in_(all_prop_ids_filtered)) | (Transaction.property_id == None)
        )
    else:
        query = query.where(Transaction.property_id.in_(all_prop_ids_filtered))

    query = query.group_by(Transaction.budget_category_id)
    result = await db.execute(query)
    results = result.all()
    
    # Map category names to their sums
    cat_actuals = {str(row[0]): float(row[1]) for row in results if row[0]}
    
    total_exec = 0.0
    total_budget_sum = 0.0

    for cat in budget.categories:
        total_budget_sum += float(cat.budgeted_amount)
        
        # Match transaction categories to budget categories using ID
        cat_actual = cat_actuals.get(str(cat.id), 0.0)
            
        cat.executed_amount = cat_actual
        total_exec += cat_actual

    budget.total_executed = total_exec
    if budget.auto_calculate_total:
        budget.total_budget = total_budget_sum
    
    # No longer committing here for performance

async def list_budgets(db: AsyncSession, property_id: Optional[str] = None, year: Optional[int] = None, month: Optional[int] = None):
    stmt = select(Budget).options(
        selectinload(Budget.categories),
        selectinload(Budget.revisions)
    )
    filters = []
    if property_id:
        filters.append(Budget.property_id == property_id)
    if year:
        filters.append(Budget.year == year)
    if month:
        filters.append(Budget.month == month)
    
    if filters:
        stmt = stmt.where(and_(*filters))
    
    result = await db.execute(stmt)
    budgets = result.scalars().all()
    for b in budgets:
        await _refresh_budget_totals(db, b)
    
    # Commit once for all refreshed budgets
    await db.commit()
    return budgets

async def create_budget(db: AsyncSession, data: BudgetCreate):
    property_id = data.property_id
    if property_id == "GENERAL":
        property_id = None

    # We create exactly 1 budget record for the specified period_type
    if data.auto_calculate_total:
        amount = sum(c.budgeted_amount for c in data.categories)
    else:
        amount = data.total_budget

    if getattr(data, 'is_annual', False):
        data.period_type = "Anual"
        data.month = 1

    new_budget = Budget(
        property_id=property_id,
        year=data.year,
        month=data.month,
        period_type=data.period_type,
        total_budget=amount,
        auto_calculate_total=data.auto_calculate_total,
        notes=data.notes
    )
    db.add(new_budget)
    await db.flush()

    for cat_data in data.categories:
        cat = BudgetCategory(
            budget_id=new_budget.id,
            category_name=cat_data.category_name,
            account_id=cat_data.account_id,
            budgeted_amount=cat_data.budgeted_amount,
            is_distributable=cat_data.is_distributable
        )
        db.add(cat)
    
    await db.commit()

    # Re-query with eager-loaded categories (db.refresh won't load relationships)
    stmt = select(Budget).options(
        selectinload(Budget.categories),
        selectinload(Budget.revisions)
    ).where(Budget.id == new_budget.id)
    result = await db.execute(stmt)
    return result.scalar_one()

async def duplicate_budget(db: AsyncSession, budget_id: str, data: BudgetDuplicate):
    source = await get_budget(db, budget_id)
    if not source:
        raise Exception("Presupuesto origen no encontrado")

    multiplier = 1 + (data.percentage_increase / 100.0)
    target_prop = None if data.target_property_id == "GENERAL" else (data.target_property_id or source.property_id)
    
    new_budget = Budget(
        property_id=target_prop,
        year=data.target_year,
        month=data.target_month,
        total_budget=float(source.total_budget) * multiplier,
        auto_calculate_total=source.auto_calculate_total,
        notes=source.notes
    )
    db.add(new_budget)
    await db.flush()

    for cat in source.categories:
        new_cat = BudgetCategory(
            budget_id=new_budget.id,
            category_name=cat.category_name,
            account_id=cat.account_id,
            budgeted_amount=float(cat.budgeted_amount) * multiplier,
            is_distributable=cat.is_distributable
        )
        db.add(new_cat)
    
    await db.commit()

    # Re-query with eager-loaded categories
    stmt = select(Budget).options(
        selectinload(Budget.categories),
        selectinload(Budget.revisions)
    ).where(Budget.id == new_budget.id)
    result = await db.execute(stmt)
    loaded = result.scalar_one()
    return [loaded]  # Return as list for compatibility

async def update_budget(db: AsyncSession, budget_id: str, data: Any, user_id: str = None): # Using Any to handle BudgetUpdate
    budget = await get_budget(db, budget_id)
    if not budget:
        return None
        
    if budget.is_closed:
        raise Exception("No se puede editar un presupuesto cerrado. Ábralo primero o duplíquelo.")
    
    old_amount = float(budget.total_budget) if not budget.auto_calculate_total else sum(float(c.budgeted_amount) for c in budget.categories)
    
    if data.notes is not None:
        budget.notes = data.notes
    
    if data.auto_calculate_total is not None:
        budget.auto_calculate_total = data.auto_calculate_total

    if data.categories is not None:
        # Complex update: replace categories
        # For simplicity, clear and re-add
        for c in budget.categories:
            await db.delete(c)
        await db.flush()
        
        for cat_data in data.categories:
            new_cat = BudgetCategory(
                budget_id=budget.id,
                category_name=cat_data.category_name,
                account_id=cat_data.account_id,
                budgeted_amount=cat_data.budgeted_amount,
                is_distributable=cat_data.is_distributable
            )
            db.add(new_cat)
        
        if data.auto_calculate_total:
            budget.total_budget = sum(c.budgeted_amount for c in data.categories)
    
    if hasattr(data, "period_type") and data.period_type is not None:
        budget.period_type = data.period_type

    # Manual total update if provided and not auto-calculating
    if data.total_budget is not None and not budget.auto_calculate_total:
        budget.total_budget = data.total_budget

    new_amount = float(budget.total_budget)
    if float(old_amount) != new_amount and user_id and hasattr(data, "justification") and data.justification:
        rev = BudgetRevision(
            budget_id=budget.id,
            user_id=user_id,
            old_amount=old_amount,
            new_amount=new_amount,
            justification=data.justification
        )
        db.add(rev)

    await db.commit()
    return await get_budget(db, budget_id)

async def get_budget(db: AsyncSession, budget_id: str) -> Optional[Budget]:
    stmt = select(Budget).options(
        selectinload(Budget.categories),
        selectinload(Budget.revisions)
    ).where(Budget.id == budget_id)
    result = await db.execute(stmt)
    budget = result.scalar_one_or_none()
    if budget:
        await _refresh_budget_totals(db, budget)
        await db.commit()
    return budget

async def calculate_distribution_keys(db: AsyncSession, parent_property_id: str) -> Dict[str, float]:
    """
    Calculates allocation percentages for sub-properties based on their active contract rent.
    """
    stmt = select(Property).where(Property.parent_id == parent_property_id)
    result = await db.execute(stmt)
    sub_units = result.scalars().all()
    
    if not sub_units:
        return {}

    allocation = {}
    total_rent = 0.0
    unit_rents = {}

    for unit in sub_units:
        contract_stmt = select(Contract).where(
            and_(
                Contract.property_id == unit.id,
                Contract.status == ContractStatus.ACTIVO.value
            )
        ).order_by(Contract.created_at.desc()).limit(1)
        
        contract_result = await db.execute(contract_stmt)
        contract = contract_result.scalar_one_or_none()
        rent = float(contract.monthly_rent) if contract else 0.0
        unit_rents[unit.id] = rent
        total_rent += rent

    if total_rent > 0:
        for unit_id, rent in unit_rents.items():
            val = float(rent) / float(total_rent)
            allocation[unit_id] = float(round(val, 4))
    else:
        count = len(sub_units)
        for unit in sub_units:
            val = 1.0 / count
            allocation[unit.id] = float(round(val, 4))

    return allocation

async def get_budget_vs_actual_report(
    db: AsyncSession, 
    property_id: str, 
    year: int, 
    month: int
) -> Dict[str, Any]:
    """
    Generates a report comparing Budgeted vs Actual amounts.
    """
    budget_stmt = select(Budget).options(selectinload(Budget.categories)).where(
        and_(
            Budget.property_id == property_id,
            Budget.year == year,
            Budget.month == month
        )
    )
    result = await db.execute(budget_stmt)
    budget = result.scalar_one_or_none()
    if not budget:
        return {"property_id": property_id, "year": year, "month": month, "rows": []}

    if budget.is_closed and budget.frozen_distribution is not None:
        dist_keys = budget.frozen_distribution
    else:
        dist_keys = await calculate_distribution_keys(db, property_id) if property_id else {}
        
    all_prop_ids = list(dist_keys.keys()) + [property_id]
    
    rows = []
    for cat in budget.categories:
        trans_stmt = select(func.sum(Transaction.amount)).where(
            and_(
                Transaction.budget_category_id == cat.id,
                Transaction.status == "Completada",
                Transaction.property_id.in_(all_prop_ids),
                _year_eq(Transaction.transaction_date, year),
                _month_eq(Transaction.transaction_date, month)
            )
        )
        result = await db.execute(trans_stmt)
        actual_total = result.scalar() or 0.0
        
        row: Dict[str, Any] = {
            "category": cat.category_name,
            "budgeted": float(cat.budgeted_amount),
            "actual": float(actual_total),
            "is_distributable": cat.is_distributable,
            "distribution": {}
        }
        
        if cat.is_distributable and dist_keys:
            for sub_id, weight in dist_keys.items():
                val = float(actual_total) * float(weight)
                row["distribution"][sub_id] = float(round(val, 2))
        
        rows.append(row)
        
    return {
        "property_id": property_id,
        "year": year,
        "month": month,
        "rows": rows
    }

async def get_budget_monthly_breakdown(db: AsyncSession, budget_id: str) -> Dict[str, Any]:
    budget = await get_budget(db, budget_id)
    if not budget:
        return None

    period = getattr(budget, "period_type", "Mensual")
    start_m = budget.month
    if period == "Bimestral": num_months = 2
    elif period == "Trimestral": num_months = 3
    elif period == "Semestral": num_months = 6
    elif period == "Anual": 
        start_m = 1
        num_months = 12
    else: num_months = 1

    months_data = []
    months_names = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    
    if budget.is_closed and budget.frozen_distribution is not None:
        dist_keys = budget.frozen_distribution
    else:
        dist_keys = await calculate_distribution_keys(db, budget.property_id) if budget.property_id else {}
        
    all_prop_ids = list(dist_keys.keys()) + [budget.property_id]
        
    from sqlalchemy import func, cast, Integer
    
    for i in range(num_months):
        target_m = start_m + i
        if target_m > 12: break
        
        m_budgeted = float(budget.total_budget) / num_months
        m_actual = 0.0
        
        m_cats = []
        for cat in budget.categories:
            c_budgeted = float(cat.budgeted_amount) / num_months

            trans_stmt = select(func.sum(Transaction.amount)).where(
                and_(
                    Transaction.budget_category_id == cat.id,
                    Transaction.direction == TransactionDirection.CREDIT.value,
                    Transaction.status == "Completada",
                    _year_eq(Transaction.transaction_date, budget.year),
                    _month_eq(Transaction.transaction_date, target_m)
                )
            )
            
            if None in all_prop_ids:
                trans_stmt = trans_stmt.where((Transaction.property_id.in_([p for p in all_prop_ids if p])) | (Transaction.property_id == None))
            else:
                trans_stmt = trans_stmt.where(Transaction.property_id.in_(all_prop_ids))
                
            res = await db.execute(trans_stmt)
            c_actual = float(res.scalar() or 0.0)
            m_actual += c_actual
            
            c_exec_pct = (c_actual / c_budgeted * 100) if c_budgeted > 0 else 0
            
            m_cats.append({
                "category_name": cat.category_name,
                "budgeted": c_budgeted,
                "actual": c_actual,
                "execution_pct": round(c_exec_pct, 2),
                "semaphore": "Verde" if c_exec_pct <= 85 else ("Amarillo" if c_exec_pct <= 100 else "Rojo")
            })
            
        m_exec_pct = (m_actual / m_budgeted * 100) if m_budgeted > 0 else 0
        
        months_data.append({
            "month": target_m,
            "month_name": months_names[target_m],
            "budgeted": m_budgeted,
            "actual": m_actual,
            "execution_pct": round(m_exec_pct, 2),
            "semaphore": "Verde" if m_exec_pct <= 85 else ("Amarillo" if m_exec_pct <= 100 else "Rojo"),
            "categories": m_cats
        })
        
    return {
        "budget_id": budget.id,
        "property_id": budget.property_id,
        "year": budget.year,
        "period_type": period,
        "total_budget": float(budget.total_budget),
        "total_actual": float(budget.total_executed),
        "execution_pct": budget.execution_pct,
        "semaphore": budget.semaphore,
        "months": months_data
    }

async def delete_budget(db: AsyncSession, budget_id: str):
    budget = await get_budget(db, budget_id)
    if budget:
        if budget.is_closed:
            raise Exception("No se puede eliminar un presupuesto que ya ha sido cerrado.")
        # Delete categories first (though they might be cascade)
        for cat in budget.categories:
            await db.delete(cat)
        await db.delete(budget)
        await db.commit()
    return True


async def export_budgets_excel(db: AsyncSession, property_id: Optional[str] = None, start_year: Optional[int] = None, end_year: Optional[int] = None):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Build query for budgets
    stmt = select(Budget).options(selectinload(Budget.categories))
    filters = []
    if property_id is not None:
        if property_id == "GENERAL" or property_id == "":
            filters.append(Budget.property_id == None)
        else:
            filters.append(Budget.property_id == property_id)
            
    if start_year:
        filters.append(Budget.year >= start_year)
    if end_year:
        filters.append(Budget.year <= end_year)
        
    if filters:
        stmt = stmt.where(and_(*filters))
        
    stmt = stmt.order_by(Budget.year.asc())
    result = await db.execute(stmt)
    budgets = result.scalars().all()
    
    for b in budgets:
        await _refresh_budget_totals(db, b)
        
    category_data = {}
    years = sorted(list(set(b.year for b in budgets)))
    if not years:
        years = [start_year] if start_year else [2026] # fallback
    
    for b in budgets:
        for cat in b.categories:
            cat_name = cat.category_name
            if cat_name not in category_data:
                category_data[cat_name] = {}
            if b.year not in category_data[cat_name]:
                category_data[cat_name][b.year] = {"budgeted": 0.0, "actual": 0.0}
            
            category_data[cat_name][b.year]["budgeted"] += float(cat.budgeted_amount)
            category_data[cat_name][b.year]["actual"] += float(cat.executed_amount)
            
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuestos"
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = ["Categoría"]
    for y in years:
        headers.extend([f"Pto. {y}", f"Ejec. {y}", f"% {y}"])
        
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    for cat_name, data_by_year in sorted(category_data.items()):
        row = [cat_name]
        for y in years:
            b_val = data_by_year.get(y, {}).get("budgeted", 0.0)
            a_val = data_by_year.get(y, {}).get("actual", 0.0)
            pct = (a_val / b_val * 100) if b_val > 0 else 0.0
            row.extend([b_val, a_val, round(pct, 2)])
        ws.append(row)
        
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for idx, cell in enumerate(row):
            if idx > 0 and (idx % 3 == 1 or idx % 3 == 2):
                cell.number_format = '"$"#,##0.00'
            elif idx > 0 and idx % 3 == 0:
                cell.number_format = '0.00"%"'
                
    ws.column_dimensions['A'].width = 30
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

async def close_budget(db: AsyncSession, budget_id: str) -> Budget:
    """Cierra un presupuesto y congela sus llaves de distribución."""
    budget = await get_budget(db, budget_id)
    if not budget:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado")
        
    if budget.is_closed:
        return budget

    # Refresh totals just in case
    await _refresh_budget_totals(db, budget)

    # Congelar distribución
    dist_keys = await calculate_distribution_keys(db, budget.property_id) if budget.property_id else {}
    budget.frozen_distribution = dist_keys
    budget.is_closed = True
    
    await db.commit()
    await db.refresh(budget)
    return budget
