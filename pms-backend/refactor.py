import sys
import os

filepath = r"c:\Users\jupa_\Desktop\proyecto real state\pms-backend\app\services\budget_service.py"
with open(filepath, "r", encoding="utf-8") as f:
    code = f.read()

# 1. Remove _ensure_general_property and GENERAL_PROPERTY_NAME
code = code.replace('GENERAL_PROPERTY_NAME = "Gastos Generales"\n', "")

s1 = """
async def _ensure_general_property(db: AsyncSession) -> str:
    \"\"\"Ensures the 'Gastos Generales' property exists and returns its ID.\"\"\"
    stmt = select(Property).where(Property.name == GENERAL_PROPERTY_NAME)
    result = await db.execute(stmt)
    prop = result.scalar_one_or_none()
    if not prop:
        # We need an owner_id. Let's find an admin or the first user.
        from app.models.user import User
        user_stmt = select(User).limit(1)
        user_result = await db.execute(user_stmt)
        user = user_result.scalar_one_or_none()
        if not user:
            raise Exception("No user found to assign Gastos Generales property")
        
        prop = Property(
            name=GENERAL_PROPERTY_NAME,
            property_type="Otros",
            address="N/A",
            city="N/A",
            latitude=0,
            longitude=0,
            area_sqm=0,
            owner_id=user.id,
            status="Disponible"
        )
        db.add(prop)
        await db.commit()
        await db.refresh(prop)
    return prop.id"""

code = code.replace(s1, "")

# 2. _refresh_budget_totals
s2 = """    dist_keys = await calculate_distribution_keys(db, budget.property_id)
    stmt_gen = select(Property).where(Property.name == GENERAL_PROPERTY_NAME).limit(1)
    gen_result = await db.execute(stmt_gen)
    gen_prop = gen_result.scalar_one_or_none()
    
    all_prop_ids = list(dist_keys.keys()) + [budget.property_id]
    if gen_prop and budget.property_id == gen_prop.id:
        all_prop_ids.append(None)"""

r2 = """    dist_keys = await calculate_distribution_keys(db, budget.property_id) if budget.property_id else {}
    all_prop_ids = list(dist_keys.keys()) + [budget.property_id]"""

code = code.replace(s2, r2)
# Also do the second occurrence in get_budget_monthly_breakdown
# Wait, s2 is unique enough. It will replace both occurrences if replace doesn't specify count.

# 3. create_budget
s3 = """    property_id = data.property_id
    if property_id == "GENERAL":
        property_id = await _ensure_general_property(db)"""

r3 = """    property_id = data.property_id
    if property_id == "GENERAL":
        property_id = None"""

code = code.replace(s3, r3)

# 4. duplicate_budget target_prop
s_target = "target_prop = data.target_property_id or source.property_id"
r_target = 'target_prop = None if data.target_property_id == "GENERAL" else (data.target_property_id or source.property_id)'
code = code.replace(s_target, r_target)

# 5. get_budget_vs_actual_report
s5 = """    dist_keys = await calculate_distribution_keys(db, property_id)
    
    # Check if this property is indeed the 'Gastos Generales' one
    stmt_gen = select(Property).where(Property.name == GENERAL_PROPERTY_NAME).limit(1)
    gen_result = await db.execute(stmt_gen)
    gen_prop = gen_result.scalar_one_or_none()
    
    all_prop_ids = list(dist_keys.keys()) + [property_id]
    if gen_prop and property_id == gen_prop.id:
        all_prop_ids.append(None) # Include transactions with NO property_id"""

r5 = """    dist_keys = await calculate_distribution_keys(db, property_id) if property_id else {}
    all_prop_ids = list(dist_keys.keys()) + [property_id]"""

code = code.replace(s5, r5)


excel_export_code = """

async def export_budgets_excel(db: AsyncSession, property_id: Optional[str] = None, start_year: Optional[int] = None, end_year: Optional[int] = None):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Build query for budgets
    stmt = select(Budget).options(selectinload(Budget.categories))
    filters = []
    if property_id is not None:
        if property_id == "GENERAL" or property_id == "":
            filters.append(Budget.property_id == None)
        else:
            filters.append(Budget.property_id == property_id)
            
    if start_year:
        filters.append(Budget.year >= start_year)
    if end_year:
        filters.append(Budget.year <= end_year)
        
    if filters:
        stmt = stmt.where(and_(*filters))
        
    stmt = stmt.order_by(Budget.year.asc())
    result = await db.execute(stmt)
    budgets = result.scalars().all()
    
    for b in budgets:
        await _refresh_budget_totals(db, b)
        
    category_data = {}
    years = sorted(list(set(b.year for b in budgets)))
    if not years:
        years = [start_year] if start_year else [2026] # fallback
    
    for b in budgets:
        for cat in b.categories:
            cat_name = cat.category_name
            if cat_name not in category_data:
                category_data[cat_name] = {}
            if b.year not in category_data[cat_name]:
                category_data[cat_name][b.year] = {"budgeted": 0.0, "actual": 0.0}
            
            category_data[cat_name][b.year]["budgeted"] += float(cat.budgeted_amount)
            category_data[cat_name][b.year]["actual"] += float(cat.executed_amount)
            
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuestos"
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = ["Categoría"]
    for y in years:
        headers.extend([f"Pto. {y}", f"Ejec. {y}", f"% {y}"])
        
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    for cat_name, data_by_year in sorted(category_data.items()):
        row = [cat_name]
        for y in years:
            b_val = data_by_year.get(y, {}).get("budgeted", 0.0)
            a_val = data_by_year.get(y, {}).get("actual", 0.0)
            pct = (a_val / b_val * 100) if b_val > 0 else 0.0
            row.extend([b_val, a_val, round(pct, 2)])
        ws.append(row)
        
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for idx, cell in enumerate(row):
            if idx > 0 and (idx % 3 == 1 or idx % 3 == 2):
                cell.number_format = '"$"#,##0.00'
            elif idx > 0 and idx % 3 == 0:
                cell.number_format = '0.00"%"'
                
    ws.column_dimensions['A'].width = 30
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
"""

code += excel_export_code

with open(filepath, "w", encoding="utf-8") as f:
    f.write(code)
print("Budget service refactored!")
